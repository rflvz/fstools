import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import os

class ExcelManager:
    def export_to_excel(self, df, output_file):
        """Export DataFrame to Excel with formatting"""
        # Generate unique filename if needed
        output_file = self._get_unique_filename(output_file)
        
        # Export to Excel
        df.to_excel(output_file, index=False)
        
        # Apply formatting
        if output_file.endswith('.xlsx'):
            self.format_excel_file(output_file)

    def format_excel_file(self, file_path):
        """Apply formatting to Excel file"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Personalizar estilos
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Aplicar formato automático a columnas
        self._auto_adjust_columns(ws)
        
        # Aplicar estilos mejorados
        self._apply_header_styles(ws, header_fill, header_font)
        self._apply_row_styles(ws, row_fill)

        wb.save(file_path)

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)  # Cap width at 50
            ws.column_dimensions[col_letter].width = adjusted_width

    def _apply_header_styles(self, ws, header_fill, header_font):
        """Apply styles to header row"""
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def _apply_row_styles(self, ws, row_fill):
        """Apply styles to data rows"""
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    def _get_unique_filename(self, file_path):
        """Generate unique filename"""
        base, ext = os.path.splitext(file_path)
        counter = 1
        new_file_path = file_path
        while os.path.exists(new_file_path):
            new_file_path = f"{base}-({counter}){ext}"
            counter += 1
        return new_file_path